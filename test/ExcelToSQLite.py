import os
import sqlite3
import openpyxl
import pandas as pd
from pathlib import Path
from dataclasses import dataclass, field

"""
NOTES:

OpenPyXL will not update formulas in rows below any deleted rows.
So, deleting blank rows is a bad idea in OpenPyXL as any
pre-existing formulas will now refer to the wrong rows.

The OpenPyXL method cell(a, b).value uses index 1,
as in (1, 1) refers to cell "A1".
"""


@dataclass
class Presupuesto:

    name: str
    file: str = field(init=False)
    processed: str = field(init=False)
    sqlite: str = field(init=False)

    def __post_init__(self):
        self.file = self.name + '.xlsx'
        self.processed = self.name + '_processed.xlsx'
        self.sqlite = Path(self.name).stem + '.db'
        self.cleaner()
        self.extractor()

    def row_eraser(self, active_ws, header_rows):

        active_ws.delete_rows(1, header_rows)

        # max_row changes in each iteration; this keeps the value fixed
        last_row = active_ws.max_row
        rows = list(active_ws.iter_rows(min_row=1, values_only=True))
        rows = reversed(rows)  # you have to start deleting from the last row

        for index, row in enumerate(rows):
            if not row[1]:
                active_ws.delete_rows(last_row - index)

    def col_eraser(self, active_ws, cols_to_delete):

        # converts the uppers to number starting with 1
        cols_to_delete = list(
            reversed([ord(column_letter) - 64 for column_letter in cols_to_delete]))

        for col in cols_to_delete:
            active_ws.delete_cols(col)

    def cleaner(self):

        wb = openpyxl.load_workbook(self.file)
        ws = wb.active

        if not ws['A'][0].value:

            # cols_to_delete = ['A', 'C', 'E', 'F', 'G', 'J', 'L', 'M', 'N']  # must be uppers, too lazy to implement upper()
            # must be uppers, too lazy to implement upper()
            cols_to_delete = ['A']

            self.row_eraser(ws, header_rows=1)
            self.col_eraser(ws, cols_to_delete)

            wb.save(self.processed)
            wb.close()

        else:
            raise Exception("Spreadsheet in the wrong format.")

    def get_depth(self, item):
        depth = item.count('.')
        return depth

    def write_parents_títulos(self, df_títulos):
        for título in df_títulos.itertuples():
            if título[0] < df_títulos.shape[0] - 1:
                parent_depth = título[4] - 1
                # query in pandas
                posible_parents = df_títulos[df_títulos.depth == parent_depth]
                posible_parents = posible_parents[posible_parents.ID < título[1]]
                if not posible_parents.empty:
                    parent_ID = posible_parents.ID.iloc[-1]
                    df_títulos.iloc[título[0], 4] = parent_ID
                else:
                    parent_ID = 0  # <<<<< CORREGIR A PARENT_ID 0
                    df_títulos.iloc[título[0], 4] = parent_ID
            else:
                df_títulos.iloc[título[0], 4] = parent_ID
        df_títulos['parent'] = df_títulos['parent'].astype(int)
        return df_títulos

    def write_parents_partidas(self, df_partidas, df_títulos):
        for partida in df_partidas.itertuples():
            parent_original_location = partida[3] - 1
            previous_original_location = df_partidas.iloc[partida[0] - 1, 2]
            if partida[0] == 0:
                parent_query = df_títulos[df_títulos.original_location ==
                                          parent_original_location]
                parent = parent_query.ID.iloc[0]
                df_partidas.iloc[partida[0], 4] = parent
            elif not previous_original_location == parent_original_location:
                parent_query = df_títulos[df_títulos.original_location ==
                                          parent_original_location]
                parent = parent_query.ID.iloc[0]
                df_partidas.iloc[partida[0], 4] = parent
            else:
                df_partidas.iloc[partida[0], 4] = parent
        df_partidas['parent'] = df_partidas['parent'].astype(int)
        return df_partidas

    def data_extractor(self, ws):
        títulos = [[0, 'PRESPUESTO_TOTAL', 0, 0, None]]
        partidas = []
        títulos_ID = 0
        partidas_ID = 0
        original_location = 0
        # iter_rows uses index 1
        for row in ws.iter_rows(values_only=True):
            original_location = original_location + 1
            if row[2]:
                name = row[1].strip()
                depth = self.get_depth(row[0])
                partidas_ID = partidas_ID + 1
                partidas.append(
                    [partidas_ID, name, original_location, depth, None])
            else:
                name = row[1].strip()
                depth = self.get_depth(row[0])
                títulos_ID = títulos_ID + 1
                títulos.append(
                    [títulos_ID, name, original_location, depth, None])
        return títulos, partidas

    def update_db(self, db_path, df_títulos, df_partidas):

        if os.path.exists(db_path):
            os.remove(db_path)

        try:
            with sqlite3.connect(db_path) as conn:
                conn.setconfig(1002)
                cursor = conn.cursor()

                # Create Títulos table
                títulos_schema = """
                    CREATE TABLE "Títulos" (
                        "ID" INTEGER PRIMARY KEY,
                        "name" TEXT,
                        "original_location" INTEGER,
                        "depth" INTEGER,
                        "parent" INTEGER,
                        FOREIGN KEY ("parent") REFERENCES "Títulos" ("ID")
                    );
                """
                cursor.execute(títulos_schema)
                df_títulos.to_sql(
                    'Títulos', conn, if_exists='append', index=False)

                # Create Partidas table
                partidas_schema = """
                    CREATE TABLE "Partidas" (
    	                "ID" INTEGER PRIMARY KEY,
    	                "name" TEXT,
    	                "original_location" INTEGER,
    	                "depth" INTEGER,
    	                "parent" INTEGER,
    	                FOREIGN KEY ("parent") REFERENCES "Títulos" ("ID")
                    );
                """
                cursor.execute(partidas_schema)
                df_partidas.to_sql(
                    'Partidas', conn, if_exists='append', index=False)

                conn.commit()
        except sqlite3.Error as e:
            print("\n" + f"SQLite error: {e}")

    def extractor(self):

        input_file = self.processed

        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        # ws.delete_rows(1, 1)

        títulos, partidas = self.data_extractor(ws)

        wb.close()

        df_títulos = pd.DataFrame(
            títulos,
            columns=['ID', 'name', 'original_location', 'depth', 'parent']
        )
        df_partidas = pd.DataFrame(
            partidas,
            columns=['ID', 'name', 'original_location', 'depth', 'parent']
        )
        df_títulos = self.write_parents_títulos(df_títulos)
        df_partidas = self.write_parents_partidas(df_partidas, df_títulos)

        db_path = f'database/{self.sqlite}'
        self.update_db(db_path, df_títulos, df_partidas)


if __name__ == '__main__':
    excel_file = input('Excel file: ')
    ppto = Presupuesto(excel_file)
