import openpyxl
from tqdm import tqdm


from pathlib import Path
from dataclasses import dataclass, field
from tkinter.filedialog import askopenfilename


"""
WARNING: the "cell()" method uses index 1
e.g. ws.cell(1,1).value refers to the content of cell A1
"""


@dataclass
class AnalisisPreciosUnitarios:
    path: str
    input: Path = field(init=False)
    output: Path = field(init=False)

    def __post_init__(self):
        self.input = Path(self.path)
        self.output = Path(self.path).with_name(
            self.input.stem + '_output.xlsx')

    def cleaner(self):

        wb = openpyxl.load_workbook(self.input)
        ws = wb.active

        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active

        last_row = ws.max_row

        headers = [
            "Item",
            "Partida",
            "Rendimiento",
            "Und_Rend",
            "Tipo_Recurso",
            "eliminar",
            "Recurso",
            "eliminar",
            "eliminar",
            "Und_Recurso",
            "Cuadrilla",
            "Cantidad",
            "Precio_Unit",
            "Precio_Parcial",
        ]

        new_sheet.append(headers)

        recursos_partidas = []
        item = 0

        listado_recurso = ("mano de obra", "materiales", "equipos")

        # iter_rows uses index 1
        for index, row in tqdm(
            enumerate(ws.iter_rows(values_only=True)),
            total=last_row,
        ):

            # rows as a list use index 0
            if row[0] and "partida" in str(row[0]).lower():

                item = item + 1
                titulo = ws.cell(index + 1, 5).value
                rendimiento = ws.cell(index + 3, 4).value
                unidad_rendimiento = ws.cell(index + 3, 3).value

                i = index + 2

                while not ("partida" in str(ws.cell(i, 1).value).lower()):

                    if i == last_row:

                        break

                    if str(ws.cell(i, 4).value).lower() in listado_recurso:

                        tipo_recurso = str(ws.cell(i, 4).value)

                    elif (
                        ws.cell(i, 3).value
                        and not ("rendimiento" in str(ws.cell(i, 1).value).lower())
                        and not ("código" in str(ws.cell(i, 1).value).lower())
                    ):

                        valores = [item, titulo,
                                   rendimiento, unidad_rendimiento]
                        valores.append(str(tipo_recurso).upper())
                        row_values = list(ws.values)[i - 1]
                        valores.extend(row_values)
                        recursos_partidas.append(valores)
                    i = i + 1

                for recursos in recursos_partidas:

                    new_sheet.append(recursos)

                recursos_partidas = []
                valores = []

        # new_sheet.delete_cols(9)
        # new_sheet.delete_cols(8)
        # new_sheet.delete_cols(6)
        new_wb.save(self.output)
        wb.close()
        new_wb.close()


if __name__ == '__main__':
    AnalisisPreciosUnitarios(askopenfilename()).cleaner()
