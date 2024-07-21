import openpyxl
from tqdm import tqdm

"""
NOTE:
OpenPyXL will not update formulas in rows below any deleted rows.
So, deleting blank rows is a bad idea in OpenPyXL as any
pre-existing formulas will now refer to the wrong rows.
"""


def active_sheet_cleaner(ws, cols_to_delete, backwards):

    empty_counter = 0
    data_counter = 0
    error_counter = 0

    last_row = ws.max_row  # max_row changes in each iteration; this keeps the value fixed
    rows = list(ws.iter_rows(min_row=1, max_col=16, values_only=True))
    if backwards == True:
        rows = reversed(rows)

    for index, row in tqdm(enumerate(rows), total=last_row):
        if not row[1]:
            empty_counter = empty_counter + 1
            ws.delete_rows(last_row - index)
        elif row[1]:
            data_counter = data_counter + 1
        else:
            error_counter = error_counter + 1

    cols_to_delete = list(reversed([ord(column_letter) - 64 for column_letter in cols_to_delete]))  # converts the uppers to number starting with 1

    ws.delete_rows(1, 6)
    for col in cols_to_delete:
        ws.delete_cols(col)

    return empty_counter, data_counter, error_counter


def main():

    file_name = input('Enter file name: ')
    input_file = file_name + '.xlsx'
    output_file = file_name + '_output.xlsx'
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    cols_to_delete = ['A', 'C', 'E', 'F', 'G', 'J', 'L', 'M', 'N', 'P']  # must be uppers, too lazy to implement upper()
    empty_rows, rows_with_data, rows_not_counted = active_sheet_cleaner(
        ws,
        cols_to_delete,
        backwards=True
    )

    wb.save(output_file)
    wb.close()

    result = f"""

    Result:
    Deleted rows: {empty_rows}
    Rows with data: {rows_with_data}
    Rows not counted: {rows_not_counted}
    """
    print(result)


if __name__ == '__main__':
    main()