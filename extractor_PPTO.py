import os
import sqlite3
import openpyxl
import pandas as pd
from tqdm import tqdm


def get_depth(item):
    depth = item.count('.')
    return depth


def write_parents_títulos(df_títulos):
    for título in tqdm(
        df_títulos.itertuples(),
        ncols=0,
        desc='Processing títulos: '
    ):
        if título[0] < df_títulos.shape[0] - 1:
            parent_depth = título[4] - 1
            posible_parents = df_títulos[df_títulos.depth == parent_depth]  # query in pandas
            posible_parents = posible_parents[posible_parents.títulos_ID < título[1]]
            if not posible_parents.empty:
                parent_ID = posible_parents.títulos_ID.iloc[-1]
                df_títulos.iloc[título[0], 4] = parent_ID
            else:
                parent_ID = 0 #### <<<<< CORREGIR A PARENT_ID 0
                df_títulos.iloc[título[0], 4] = parent_ID
        else:
            df_títulos.iloc[título[0], 4] = parent_ID
    df_títulos['parent'] = df_títulos['parent'].astype(int)
    return df_títulos


def write_parents_partidas(df_partidas, df_títulos):
    for partida in tqdm(
        df_partidas.itertuples(),
        ncols=0,
        desc='Processing partidas: '
    ):
        parent_original_location = partida[3] - 1
        previous_original_location = df_partidas.iloc[partida[0] - 1, 2]
        if partida[0] == 0:
            parent_query = df_títulos[df_títulos.original_location == parent_original_location]
            parent = parent_query.títulos_ID.iloc[0]
            df_partidas.iloc[partida[0], 4] = parent
        elif not previous_original_location == parent_original_location:
            parent_query = df_títulos[df_títulos.original_location == parent_original_location]
            parent = parent_query.títulos_ID.iloc[0]
            df_partidas.iloc[partida[0], 4] = parent
        else:
            df_partidas.iloc[partida[0], 4] = parent
    df_partidas['parent'] = df_partidas['parent'].astype(int)
    return df_partidas


def data_extractor(ws):
    títulos = [[0, 'PLACEHOLDER', 0, 0, None]]
    partidas = []
    títulos_ID = 0
    partidas_ID = 0
    original_location = 0
    # iter_rows uses index 1
    for index, row in tqdm(
        enumerate(ws.iter_rows(min_row=1, max_col=6, values_only=True)),
        total=ws.max_row,
        ncols=0,
        desc='Extracting from xlsx: ',
    ):
        original_location = original_location + 1
        if row[2]:
            name = row[1].strip()
            depth = get_depth(row[0])
            partidas_ID = partidas_ID + 1
            partidas.append([partidas_ID, name, original_location, depth, None])
        else:
            name = row[1].strip()
            depth = get_depth(row[0])
            títulos_ID = títulos_ID + 1
            títulos.append([títulos_ID, name, original_location, depth, None])
    return títulos, partidas


def update_db(db_path, df_títulos, df_partidas):

    if os.path.exists(db_path):
        os.remove(db_path)

    try:
        with sqlite3.connect(db_path) as conn:
            conn.setconfig(1002)
            cursor = conn.cursor()

            # Create Títulos table
            títulos_schema = """
                CREATE TABLE Títulos (
                    "títulos_ID" INTEGER PRIMARY KEY,
                    "name" TEXT,
                    "original_location" INTEGER,
                    "depth" INTEGER,
                    "parent" INTEGER,
                    FOREIGN KEY ("parent") REFERENCES "Títulos" ("títulos_ID")
                );
            """
            cursor.execute(títulos_schema)
            df_títulos.to_sql('Títulos', conn, if_exists='append', index=False)

            # Create Partidas table
            partidas_schema = """
                CREATE TABLE "Partidas" (
	                "partidas_ID" INTEGER PRIMARY KEY,
	                "name" TEXT,
	                "original_location" INTEGER,
	                "depth" INTEGER,
	                "parent" INTEGER,
	                FOREIGN KEY ("parent") REFERENCES "Títulos" ("títulos_ID")
                );
            """
            cursor.execute(partidas_schema)
            df_partidas.to_sql('Partidas', conn, if_exists='append', index=False)

            conn.commit()
    except sqlite3.Error as e:
        print("\n" + f"SQLite error: {e}")
    finally:
            cursor.close()


def main():
    file_name = input('Enter file name: ')
    input_file = file_name + '.xlsx'

    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    ws.delete_rows(1, 1)

    títulos, partidas = data_extractor(ws)

    wb.close()

    df_títulos = pd.DataFrame(
        títulos,
        columns=['títulos_ID', 'name', 'original_location', 'depth', 'parent']
    )
    df_partidas = pd.DataFrame(
        partidas,
        columns=['partidas_ID', 'name', 'original_location', 'depth', 'parent']
    )
    df_títulos = write_parents_títulos(df_títulos)
    df_partidas = write_parents_partidas(df_partidas, df_títulos)

    db_path = f'database/{file_name}.db'
    update_db(db_path, df_títulos, df_partidas)

    print('Sript completed succesfully!')

if __name__ == '__main__':
    main()