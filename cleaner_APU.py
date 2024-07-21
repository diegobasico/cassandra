import openpyxl
from tqdm import tqdm


def main():

    input_file = r'C:/Users/Diego/Desktop/APU.xlsx'
    output_file = r'C:/Users/Diego/Desktop/output_ie.xlsx'

    # input_file = input('Enter input xlsx: ')
    # output_file = input('Enter output xlsx: ')

    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    last_row = ws.max_row

    headers = [
        "Item",
        "Partida",
        "Rendimiento",
        "Und_Rend",
        "Tipo_Recurso",
        "eliminar",
        "Recurso",
        "eliminar",
        "eliminar",
        "Und_Recurso",
        "Cuadrilla",
        "Cantidad",
        "Precio_Unit",
        "Precio_Parcial",
    ]

    new_sheet.append(headers)

    recursos_partidas = []
    item = 0

    listado_recurso = ("mano de obra", "materiales", "equipos")

    # iter_rows uses index 1
    for index, row in tqdm(
        enumerate(ws.iter_rows(min_row=1, max_col=9, values_only=True)),
        total=last_row,
    ):

        # rows as a list use index 0
        if row[0] and "partida" in str(row[0]).lower():

            item = item + 1  # cells use index 1
            titulo = ws.cell(index + 1, 4).value
            rendimiento = ws.cell(index + 3, 3).value
            unidad_rendimiento = ws.cell(index + 3, 2).value

            i = index + 2

            while not ("partida" in str(ws.cell(i, 1).value).lower()):

                if i == last_row:

                    break

                if str(ws.cell(i, 3).value).lower() in listado_recurso:

                    tipo_recurso = str(ws.cell(i, 3).value)

                elif (
                    ws.cell(i, 2).value
                    and not ("rendimiento" in str(ws.cell(i, 1).value).lower())
                    and not ("código" in str(ws.cell(i, 1).value).lower())
                ):

                    valores = [item, titulo, rendimiento, unidad_rendimiento]
                    valores.append(str(tipo_recurso).upper())
                    row_values = list(ws.values)[i - 1]
                    valores.extend(row_values)
                    recursos_partidas.append(valores)
                i = i + 1

            for recursos in recursos_partidas:

                new_sheet.append(recursos)

            recursos_partidas = []
            valores = []

    new_sheet.delete_cols(9)
    new_sheet.delete_cols(8)
    new_sheet.delete_cols(6)
    new_wb.save(output_file)
    wb.close()
    new_wb.close()


if __name__ == '__main__':
    main()